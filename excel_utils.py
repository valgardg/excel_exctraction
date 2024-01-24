from config import EXCEL_FILE_PATH
import openpyxl

# returns workbook object
def get_workbook(file_path):
    return openpyxl.load_workbook(file_path)

# returns sheet with name sheet_name
def get_sheet(sheet_name):
    workbook = get_workbook(EXCEL_FILE_PATH)
    if(sheet_name not in workbook.sheetnames):
        raise Exception("Sheet name not found")
    return workbook[sheet_name]

# returns value of cell at row, col
def read_cell(sheet, row, col):
    return sheet.cell(row=row, column=col).value

# return data from a sheet with given sheet_name, sheet_dimensions, header_strings
def get_data_from_sheet(sheet_name, sheet_dimensions, header_strings):
    # fetch plant sheet
    sheet = get_sheet(sheet_name, )
    # read plants
    data = []
    for row in range(sheet_dimensions[0], sheet_dimensions[2]+1):
        data_info = {}
        for col in range(sheet_dimensions[1], sheet_dimensions[3]+1):
            if(read_cell(sheet, row, col) == None):
                continue
            mapped_key = header_strings.get(col, col)
            data_info[mapped_key] = read_cell(sheet, row, col)
        if(len(data_info) > 0):
            data.append(data_info)
    return data